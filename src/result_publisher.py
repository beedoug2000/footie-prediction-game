import argparse
from openpyxl import load_workbook
# import sys
import logging
import shutil

### Set up the logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
s_handler = logging.StreamHandler()
s_formatter = logging.Formatter('%(message)s')
s_handler.setFormatter(s_formatter)
logger.addHandler(s_handler)

### Start of processing
logger.info("Preparing file for publication!")

### Get the filename from the command line and open the file for processing
parser = argparse.ArgumentParser(description='Prepare tournament results file for publication', usage='result_publisher.py <file_name> <Optional: --test>')
parser.add_argument('file_name', help='The name of the file to process')
# parser.add_argument('target_dir', help='The name of the directory in which to store the public result file')
parser.add_argument('--test', help='If specified, the script will run in test mode', action='store_true')
args = parser.parse_args()

filename = args.file_name
filename_parts = filename.split('.')
# target_dir = args.target_dir
directory = filename_parts[0].split('/')[0]


new_filename = None

### If in test mode, create a copy of the specified file and use that
if(args.test):
    new_filename = "{}_Test.{}".format(filename_parts[0], filename_parts[1])
    logger.info("Running in test mode.")
else:
    new_filename = "{}/index.{}".format(directory, filename_parts[1])
    # new_filename = "{}_Public.{}".format(filename_parts[0], filename_parts[1])

shutil.copyfile(filename, new_filename)
filename = new_filename
logger.info("Reading file {}".format(filename))

file = None

try:
    file = load_workbook(filename=filename)
except FileNotFoundError:
    raise SystemExit("No such file or directory: '{}'".format(filename))

### Get all the players on the Leaderboard and rename them
logger.info("Renaming players on Leaderboard")

leader_sheet = file['Leaderboard']
row_number = 3
while(leader_sheet.cell(row=row_number, column=2).value):
    player_name = leader_sheet.cell(row=row_number, column=2).value
    logger.debug("Renaming {} to {}".format(player_name, player_name[0]))
    leader_sheet.cell(row=row_number, column=2).value = player_name[0]
    row_number += 1
file.save(filename=filename)

### Get all the sheets and rename them
logger.info("Renaming player sheets")

sheetnames = file.sheetnames
for sheetname in sheetnames:
    if sheetname in ('Matches', 'Leaderboard'):
        logger.debug("'{}' not a player sheet - continuing.".format(sheetname))
        continue
    
    new_sheetname = sheetname[0]*7
    logger.debug("Renaming sheet {} to {}".format(sheetname, new_sheetname))
    file[sheetname].title = new_sheetname
    
logger.info("Finished renaming all players")
file.save(filename=filename)