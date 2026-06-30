import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color
import logging
import shutil
import os

### Set up the logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
s_handler = logging.StreamHandler()
s_formatter = logging.Formatter('%(message)s')
s_handler.setFormatter(s_formatter)
logger.addHandler(s_handler)

### Start of processing
logger.info("Starting the update")

### Get the filename from the command line and open the file for processing
parser = argparse.ArgumentParser(description='Update knock-out stage bracket with team names.', usage='{} <file_name> <Optional: --test> <list_of_replacements>'.format(os.path.basename(__file__)))
parser.add_argument('file_name', help='The name of the file to process')
parser.add_argument('--test', help='If specified, the script will run in test mode', action='store_true')
parser.add_argument('list_of_replacements', help='A string representing a list of names to replace in the knock-out bracket in the format (<name_to_replace>, <replacement_name>, ...)')
args = parser.parse_args()

filename = args.file_name

### If in test mode, create a copy of the specified file and use that
if(args.test):
    filename_parts = filename.split('.')
    test_filename = "{}_Test.{}".format(filename_parts[0], filename_parts[1])
    shutil.copyfile(filename, test_filename)
    filename = test_filename
    logger.info("Running in test mode.")

list_of_replacements = args.list_of_replacements.split(",")
logger.info("Updating file {}".format(filename))

replacement_names = {}
for i in range(0, len(list_of_replacements), 2):
    replacement_names[list_of_replacements[i].strip()] = list_of_replacements[i+1].strip()

file = None

try:
    file = load_workbook(filename=filename)
except FileNotFoundError:
    raise SystemExit("No such file or directory: '{}'".format(filename))

logger.debug("Reading team cells from file")
matches_sheet = file['Matches']
matches_row_sets = [(5,6,7,8), (11,12,13,14), (17,18,19,20)]

teams = {}

for col in range(3, 16, 4):
    for matches_row_set in matches_row_sets:
        for matches_row in matches_row_set:
            # team_cell = matches_sheet.cell(row=matches_row, column=col)
            team_name = matches_sheet.cell(row=matches_row, column=col).value
            teams[team_name] = [matches_row, col]

sheetnames = file.sheetnames
for sheetname in sheetnames:
    if sheetname in ('Leaderboard'):
        logger.debug("'{}' not a knock-out stage sheet - continuing.".format(sheetname))
        continue
    
    logger.info("Updating {}".format(sheetname))
    for col in file[sheetname].iter_cols(min_row=64, max_row=110, min_col=3, max_col=19):
        for cell in col:
            if replacement_names.get(cell.value):
                logger.info("    Replacing {} with {}".format(cell.value, replacement_names[cell.value]))
                
                new_name = replacement_names[cell.value]
                target_cell = matches_sheet.cell(row=teams[new_name][0], column=teams[new_name][1])
                target_cell_colour = target_cell.fill.start_color
                
                cell.value = new_name
                cell.alignment = Alignment(horizontal="center", vertical="bottom")
                
                if target_cell_colour.type == "theme":
                    c = Color(theme=target_cell_colour.theme, tint=target_cell_colour.tint)
                    cell.fill = PatternFill(start_color=c, fill_type = "solid")
                else:
                    cell.fill = PatternFill(start_color=target_cell_colour.rgb, fill_type = "solid")
                
            
file.save(filename=filename)
