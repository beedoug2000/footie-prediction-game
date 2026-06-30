import argparse
from openpyxl import load_workbook
import sys
import logging
import shutil

### Set up the logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
s_handler = logging.StreamHandler()
s_formatter = logging.Formatter('%(message)s')
s_handler.setFormatter(s_formatter)
logger.addHandler(s_handler)

### Functions

# Gets all the group stage results from a sheet and stores them in a dictionary for later processing
def get_group_results(sheet):
    result_col_sets = [(3,4,5), (7,8,9), (11,12,13)]
    group_results = {}

    for row_num in range(27, 58):
        for result_col_set in result_col_sets:
            teamA = sheet.cell(row=row_num, column=result_col_set[0]).value
            teamB = sheet.cell(row=row_num, column=result_col_set[1]).value
            result = sheet.cell(row=row_num, column=result_col_set[2]).value

            # Only add this to the results if that result is recorded in the Matches sheet
            # This is cheating - should really check if all 3 values are present but will assume that the team names will always be populated
            if result:
                group_results[teamA.strip()+'-'+teamB.strip()] = result.strip()

    return group_results

# Gets all the results for a particular knock-out stage from a sheet and stores them in a dictionary for later processing
def get_knockout_stage_results(sheet, start_row, interval, num_times, result_col_set):
    knockout_stage_results = {}
    
    for i in range(0, num_times):
        row_num = start_row + (i * interval)
        teamA = sheet.cell(row=row_num, column=result_col_set[0]).value
        result = sheet.cell(row=row_num, column=result_col_set[1]).value
        row_num += 1
        teamB = sheet.cell(row=row_num, column=result_col_set[0]).value
        
        if result:
            knockout_stage_results[teamA.strip()+'-'+teamB.strip()] = result.strip()
    
    return knockout_stage_results

# Gets all the knock-out results from a sheet and stores them in one dictionary for later processing
def get_knockout_results(sheet):
    knockout_results = get_knockout_stage_results(sheet, 64, 3, 16, (3,4))
    knockout_results.update(get_knockout_stage_results(sheet, 65, 6, 8, (7,8)))
    knockout_results.update(get_knockout_stage_results(sheet, 68, 12, 4, (11,12)))
    knockout_results.update(get_knockout_stage_results(sheet, 74, 24, 2, (15,16)))
    knockout_results.update(get_knockout_stage_results(sheet, 84, 4, 2, (19,20)))

    return knockout_results

# Compares a predicted result with an actual result and returns the number of points earned
def compare_results(predicted_result, actual_result):        
    predicted = predicted_result.split('-')
    p_team_A = int(predicted[0].strip())
    p_team_B = int(predicted[1].strip())
    
    actual = actual_result.split('-')
    a_team_A = int(actual[0].strip())
    a_team_B = int(actual[1].strip())

    # Correct score
    if p_team_A == a_team_A and p_team_B == a_team_B:
        return 2

    # Correct result
    p_diff = p_team_A - p_team_B
    a_diff = a_team_A - a_team_B

    if p_diff == a_diff or (p_diff < 0 and a_diff < 0) or (p_diff > 0 and a_diff > 0):
        return 1

    # Completely wrong
    return 0

### Updates a sheet with group standings
def update_groups(sheet, col_range, row_sets, groups):
    for col in col_range:
        for row_set in row_sets:
            group_name = sheet.cell(row=row_set[0]-1, column=col).value
            group = groups[group_name]
            row_num = row_set[0]
            for team, points in group.items():
                sheet.cell(row=row_num, column=col).value = team
                sheet.cell(row=row_num, column=col+1).value = points
                row_num += 1

### Start of processing
logger.info("Starting processing of results!")

### Get the filename from the command line and open the file for processing
parser = argparse.ArgumentParser(description='Process tournament results against player predictions.', usage='result_processor_wc.py <file_name> <Optional: --test>')
parser.add_argument('file_name', help='The name of the file to process')
parser.add_argument('--test', help='If specified, the script will run in test mode', action='store_true')
args = parser.parse_args()

filename = args.file_name

### If in test mode, create a copy of the specified file and use that
if(args.test):
    filename_parts = filename.split('.')
    test_filename = "{}_Test.{}".format(filename_parts[0], filename_parts[1])
    shutil.copyfile(filename, test_filename)
    filename = test_filename
    logger.info("Running in test mode.")

logger.info("Reading from file {}".format(filename))

file = None

try:
    file = load_workbook(filename=filename)
except FileNotFoundError:
    raise SystemExit("No such file or directory: '{}'".format(filename))

### Get all the available group match results
matches_sheet = file['Matches']
results = get_group_results(matches_sheet)

logger.debug("")
logger.debug("LATEST GROUP STAGE RESULTS:")
logger.debug(results)
logger.debug("")

if len(results.keys()) == 0:
    sys.exit("No match results have been recorded, nothing to process.")
    
### Calculate the points of each team
logger.info("Calculating current points total of each team")
logger.debug("")
team_points = {}
for result in results.keys():
    teams = result.split('-')
    team_A = teams[0].strip()
    team_B = teams[1].strip()
    
    score = results[result].split('-')
    team_A_score = int(score[0].strip())
    team_B_score = int(score[1].strip())

    if team_B_score < team_A_score:
        logger.debug("{} beat {} {}, adding 3 points to their total.".format(team_A, team_B, results[result]))
        team_points[team_A] = team_points.get(team_A, 0)+3
        team_points[team_B] = team_points.get(team_B, 0)
    elif team_A_score < team_B_score:
        logger.debug("{} beat {} {}, adding 3 points to their total.".format(team_B, team_A, results[result][::-1]))
        team_points[team_B] = team_points.get(team_B, 0)+3
        team_points[team_A] = team_points.get(team_A, 0)
    else:
        logger.debug("{} and {} drew {}, adding 1 point to each of their totals.".format(team_A, team_B, results[result]))
        team_points[team_A] = team_points.get(team_A, 0)+1
        team_points[team_B] = team_points.get(team_B, 0)+1

logger.debug("")
logger.debug("TEAM POINTS:")
logger.debug(team_points)
logger.debug("")
        
### Get each group of teams so results can be sorted properly
logger.debug("Reading team groups from file")
groups = {}
matches_row_sets = [(5,6,7,8), (11,12,13,14), (17,18,19,20)]

for col in range(3, 16, 4):
    for matches_row_set in matches_row_sets:
        group_name = matches_sheet.cell(row=matches_row_set[0]-1, column=col).value
        group = {}
        for matches_row in matches_row_set:
            team = matches_sheet.cell(row=matches_row, column=col).value
            group[team] = 0
        groups[group_name] = group

logger.debug("")
logger.debug("GROUPS:")
logger.debug(groups)
logger.debug("")

### Update each teams' points total and sort accordingly
logger.debug("Updating points and sorting team standings by group")
for group_name, group in groups.items():
    for team in group.keys():
        group[team] = team_points.get(team, 0)
    
    sorted_group = dict(sorted(group.items(), key=lambda item: item[1], reverse=True))
    groups[group_name] = sorted_group

logger.debug("")
logger.debug("GROUPS:")
logger.debug(groups)
logger.debug("")

### Write the group standings to the Matches sheet
logger.info("Updating Matches with team standings")
update_groups(matches_sheet, range(3, 16, 4), matches_row_sets, groups)
file.save(filename=filename)

### Write the group standings to the Leaderboard sheet
logger.info("Updating Leaderboard with team standings")
leader_sheet = file['Leaderboard']
leader_row_sets = [(3,4,5,6), (9,10,11,12), (15,16,17,18)]
update_groups(leader_sheet, range(5, 15, 3), leader_row_sets, groups)
file.save(filename=filename)

### Read in the knock-out results and add to the results dictionary
# This is done after calculating the group stage points so that these results do not affect those totals
results.update(get_knockout_results(matches_sheet))

logger.debug("")
logger.debug("LATEST RESULTS INCLUDING KNOCK-OUT STAGES:")
logger.debug(results)
logger.debug("")

### Get all the player predictions
logger.info("Loading all player predictions")
player_predictions = {}
sheetnames = file.sheetnames
for sheetname in sheetnames:
    if sheetname in ('Matches', 'Leaderboard'):
        logger.debug("'{}' not a player sheet - continuing.".format(sheetname))
        continue

    predictions = get_group_results(file[sheetname])
    predictions.update(get_knockout_results(file[sheetname]))
    
    if len(predictions.keys()) != 0:
        player_predictions[sheetname] = predictions
    
    update_groups(file[sheetname], range(3, 16, 4), matches_row_sets, groups)

logger.debug("")
logger.debug("PLAYER PREDICTIONS:")
logger.debug(player_predictions)
logger.debug("")

player_points = {}

### Compare player predictions to match results and total up scores
for player in player_predictions.keys():
    logger.info("Processing {}'s predictions".format(player))
    
    total_points = 0
    for match in results:
        points = 0
        
        predicted_result = player_predictions[player].get(match)

        if predicted_result:
            actual_result = results[match]

            logger.debug("Comparing results for {}".format(match))
            logger.debug("Predicted: {}".format(predicted_result))
            logger.debug("Actual: {}".format(actual_result))

            if "-" not in predicted_result:
                sys.exit("Predicted result of {} for {} by player {} in unknown format (expecting format number-number, as in 2-1).  Unsure how to proceed.  Exiting...".format(predicted_result, match, player))
            
            points = compare_results(predicted_result, actual_result)
            logger.debug("Points earned on this result: {}".format(points))
        else:
            logger.debug("No prediction entered for {} for {} - 0 points earned.".format(player, match))
            
        total_points += points

    logger.debug("Total points earned: {}".format(total_points))
    logger.debug("")
    player_points[player] = total_points

### Write players to the leaderboard in order
logger.info("Updating Leaderboard with player standings")
sorted_player_points = dict(sorted(player_points.items(), key=lambda item: item[1], reverse=True)) 

row_number = 3
for player in sorted_player_points.keys():
    s = 's' if sorted_player_points[player] > 1 else ''
    logger.info("Adding player {} with {} point{} to Leaderboard".format(player, sorted_player_points[player], s))
    leader_sheet.cell(row=row_number, column=2).value = player
    leader_sheet.cell(row=row_number, column=3).value = sorted_player_points[player]
    row_number += 1

# In theory, there should always be the same number of players playing
# But just in case someone dropped out, clear all cells beyond the latest leaderboard
while(leader_sheet.cell(row=row_number, column=2).value or leader_sheet.cell(row=row_number, column=3).value):
    leader_sheet.cell(row=row_number, column=2).value = None
    leader_sheet.cell(row=row_number, column=3).value = None
    row_number += 1

file.save(filename=filename)